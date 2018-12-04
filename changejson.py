# -*- coding: utf-8 -*-
"""
Created on Wed Nov 28 21:06:43 2018

@author: Ren_xc
"""

from openpyxl import load_workbook
import pandas as pd


# load the excle workbook
wb = load_workbook('../data/site_list.xlsx')
# get the sheets' name
sheetnames = wb.get_sheet_names()  
# choose the sheet you want
ws = wb.get_sheet_by_name(sheetnames[0])
# define lists to save site longitude and latitude
site,lon,lat = [],[],[]
# read the value to lists
for cell in ws['B']:
    site.append(cell.value)
for cell in ws['D']:
    lon.append(cell.value)
for cell in ws['E']:
    lat.append(cell.value)
# delete the title in first one
del site[0],lon[0],lat[0]

inputfile = '../data/site_20140513-20180908/pm.csv'

readdata = pd.read_csv(inputfile)
data = readdata.drop(['date'], axis = 1)
clname = data.columns.values.tolist()
for n in clname:
    data[n].fillna(data[n].mean(),inplace = True)
value = list(data.loc[0])

del site[26], lon[26], lat[26], value[26]

file = open(r'../data/point.json','w')
#for i in range(len(lat)):
#    str_temp = '{"lat":' + str(lat[i]) + ',"lng":' + str(lon[i]) + ',"count":' + str(value[i]) +'},\n'
#    file.write(str_temp)
for i in range(len(lat)):
    str_temp = "{name: '" + str(site[i]) + "', value:" + str(value[i]) +"},\n"
    file.write(str_temp)
for i in range(len(lat)):
    str_lonlat = "'" + str(site[i]) + "':[" + str(lon[i]) + "," + str(lat[i]) + "],\n"
    file.write(str_lonlat)    
file.close()